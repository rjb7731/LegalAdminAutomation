from openpyxl import load_workbook
import pyautogui
import pygetwindow
import time
import pyperclip
import sys

window = pygetwindow.getWindowsWithTitle('WINDOW NAME HERE')[0]

if window.isMaximized == False:
    window.maximize()

workbook = load_workbook(filename="EXCEL FILE WITH INFO GOES HERE")
sheet = workbook.active

FirstRow = []
SecondRow = []
Mattername = []

matnum = str(sheet.max_row)

time.sleep(2)

for row in sheet['A1':'A'+matnum]:  # Set First Entry from spreadsheet
    for cell in row:
        FirstRow.append(cell.value)

for row in sheet['B1':'B'+matnum]:  # Set Second Data Entry from spreadsheet
    for cell in row:
        SecondRow.append(cell.value)

for row in sheet['C1':'C'+matnum]:  # Set Second Data Entry from spreadsheet
    for cell in row:
        Mattername.append(cell.value)

def loadmatter():
    """enter entries and load matter"""
    pyautogui.moveTo(408, 199)  # move to first position
    pyautogui.click()
    pyautogui.typewrite(str(FirstRow[entry]))
    pyautogui.moveTo(707, 200)  # move to second position
    pyautogui.click()
    pyautogui.typewrite(str(SecondRow[entry]))
    pyautogui.moveTo(1258, 209)  # moves to load button
    pyautogui.click()

def datecheck():
    """check date"""
    pyautogui.moveTo(1250,345) #date
    pyautogui.click()
    pyautogui.moveTo(1113,848) #edit
    pyautogui.click()
    time.sleep(0.2)
    found = copy_clipboard(1283,517) #actual date grab
    time.sleep(0.2)
    pyautogui.click(1428,842) #cancel button
    pyautogui.click()
    time.sleep(0.8)
    pyautogui.click(1103,353)
    pyautogui.click()
    time.sleep(0.2)
    return found

def fire():
    """destroy file"""
    pyautogui.moveTo(1544, 258)  # moves to fire button
    pyautogui.click()
    pyautogui.moveTo(979, 581)  # moves to yes button
    time.sleep(1)
    pyautogui.click()
    time.sleep(1.5)
    pyautogui.moveTo(1158, 579)
    time.sleep(2)
    pyautogui.click()
    time.sleep(1)
    

def clearandback():
    """move back to start and clear entries"""
    pyautogui.moveTo(405, 202)  # moves back to first box to clear
    pyautogui.doubleClick()
    time.sleep(0.5)
    pyautogui.press('delete')
    pyautogui.moveTo(707, 201)  # moves to second box to clear
    pyautogui.doubleClick()
    time.sleep(0.5)
    pyautogui.press('delete')
        
        
def copy_clipboard(x, y):
    """copy to clipboard"""
    pyperclip.copy("")# <- This empties the clipboard
    pyautogui.moveTo(x, y) #move to position to check
    time.sleep(0.2)
    pyautogui.click()
    pyautogui.keyDown('shift')
    pyautogui.press('end')
    time.sleep(0.2)
    pyautogui.keyUp('shift')
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.2) 
    return pyperclip.paste()

def checkdescription():
    """move to description to check"""
    time.sleep(0.2)
    found = copy_clipboard(1169, 393)
    sheetfound = f'{Mattername[entry]}'
    if found == sheetfound:
        return "Match"
    else:
        pyautogui.alert(f"ERROR: These Do Not Match, \n \n Spreadsheet Entry = {sheetfound}, \n \n Paperclip Found = {found}")
        sys.exit("ERROR: Sheet and screen do not match")

for entry in range(1, int(matnum)):  # set iterations
    loadmatter()
    time.sleep(0.8)
    match = checkdescription()
    date = datecheck()
    if match in "Match":
        print(f"{FirstRow[entry]},{SecondRow[entry]} -DATE FOUND: {date}")      
    else:
        sys.exit("ERROR - Please try again - (**matters did not match**)")
    if '2020'in date:
            pass
    else:
            sys.exit(f"{FirstRow[entry]},{SecondRow[entry]}- {date} -(**BAD DATE**)")
    fire()
    print("/\DESTROYED")
    time.sleep(2.5)
    clearandback()
    
